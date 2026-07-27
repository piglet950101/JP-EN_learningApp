"""
Excel -> JSON importer for 快単 (Kaitan).

Source: ../../Appli開発［foxgold共有］/アプリ基本データ-完成版.xlsx (sheet: 快単vol.1)
Output: assets/content/words_vol1.json

Schema (live, 9 cols):
  A: B          (block-marker, value only at first row of each block — informational)
  B: 見出し№   (word id, str like "0001")
  C: 英単語     (headword)
  D: <no hdr>  品詞   (single value; may carry multi-meaning marker like "形(2)" or "名2")
  E: 意味       (meaning(s); may be comma/、-separated)
  F: 覚え方     (mnemonic type, possibly combined like "ゴロ・語源")
  G: 覚え方の具体的方法 (mnemonic text; types separated by ／)
  H: 例文       (English example)
  I: 日本語訳   (JP translation)

Rules (client-confirmed):
- POS multi-meaning:
    "名"     -> single meaning
    "名2"    -> two meanings, BOTH required to answer    (mode: both_required)
    "名(2)"  -> two meanings, EITHER acceptable          (mode: either_ok)
    (any base POS may carry the 2 / (2) suffix)
- Block derivation (do NOT trust any sheet column for this):
    vol.1: block = ceil(id / 48)
    vol.2: block = 23 + ceil((id - 1100) / 48)
- Mnemonic combo: F joined by "・", G correspondingly split by "／".
"""

from __future__ import annotations
import json, math, re, sys, os
from pathlib import Path
import openpyxl
from openpyxl.cell.rich_text import CellRichText

ROOT = Path(__file__).resolve().parents[2]           # …/20260521_english_app
SRC_DIR = ROOT / "Appli開発［foxgold共有］"
OUT  = Path(__file__).resolve().parents[1] / "assets" / "content" / "words.json"


def _pick_newest(patterns: list[str]) -> Path | None:
    """Return the newest file matching any of the patterns, by mtime."""
    cands: list[Path] = []
    for p in patterns:
        cands += list(SRC_DIR.glob(p))
    if not cands:
        return None
    return max(cands, key=lambda p: p.stat().st_mtime)


def _pick_vol1_source() -> Path:
    """Vol.1 masters historically: 'アプリ基本データ-完成版*.xlsx',
       optionally with co-director date suffix and Windows '(1)' duplicates."""
    p = _pick_newest([
        "アプリ基本データ-完成版*.xlsx",
        "アプリ基本データ完成版*.xlsx",        # without the hyphen
        "Apuri基本データ-完成版*.xlsx",         # alt romanization
        "Apuri基本データ完成版*.xlsx",
    ])
    if p is None:
        print(f"No vol.1 master Excel found under {SRC_DIR}", file=sys.stderr)
        sys.exit(1)
    # Reject any that look explicitly like vol.2 (just in case glob is too broad).
    if "vol.2" in p.name or "_2_" in p.name:
        # Try to find another candidate that isn't vol.2.
        alts = [c for c in SRC_DIR.glob("*完成版*.xlsx")
                if "vol.2" not in c.name and "_2_" not in c.name]
        if alts:
            p = max(alts, key=lambda x: x.stat().st_mtime)
    return p


def _pick_vol2_source() -> Path | None:
    """Vol.2 master: anything matching '*vol.2*' or '*1101-2201*'. Optional —
       returns None if vol.2 data hasn't been delivered yet."""
    return _pick_newest([
        "*vol.2*.xlsx",
        "*1101-2201*.xlsx",
    ])

POS_BASE = {"他","自","名","形","副","前","接","間"}

POS_RE = re.compile(r"^(?P<base>[他自名形副前接間]+)(?P<paren>\(?\d?\)?)$")

def parse_pos(raw: str):
    """Return (pos_list, meaning_mode) where pos_list is the base POS chars,
    meaning_mode in {'single','both_required','either_ok'}."""
    if not raw:
        return [], "single"
    s = raw.strip()
    # Detect either-ok marker: "(2)" anywhere -> either_ok
    if "(2)" in s or "(2)" in s.replace("（","(").replace("）",")"):
        s_clean = s.replace("(2)","").replace("（2）","").strip()
        base = [c for c in s_clean if c in POS_BASE]
        return base or [s_clean], "either_ok"
    # Detect both-required marker: bare "2" suffix -> both_required
    if s.endswith("2") and not s.endswith("(2)"):
        s_clean = s[:-1].strip()
        base = [c for c in s_clean if c in POS_BASE]
        return base or [s_clean], "both_required"
    base = [c for c in s if c in POS_BASE]
    return base or [s], "single"

def split_meanings(meaning: str, mode: str, pos_count: int = 1):
    """Split the 意味 column into a list of meanings.

    Separator depends on context (observed in the live data):
      • single POS + 「2」 / 「(2)」 marker → comma 「、」 separator
      • multi POS                          → centered-dot 「・」 separator
                                            (one meaning per POS, paired by index)
    """
    if not meaning:
        return []
    if pos_count >= 2:
        # Multi-POS: split on ・ (and fall back to 、 if no ・ present).
        parts = [p.strip() for p in re.split(r"[・･]", meaning) if p.strip()]
        if len(parts) > 1:
            return parts
        # Fallback to comma split if no centered dot.
        parts = [p.strip() for p in re.split(r"[、,]", meaning) if p.strip()]
        return parts or [meaning.strip()]
    if mode == "single":
        return [meaning.strip()]
    parts = [p.strip() for p in re.split(r"[、,]", meaning) if p.strip()]
    return parts or [meaning.strip()]

# Normalization for the messy abbreviations the client uses in F column.
MNEMONIC_NORMALIZE = {
    "ごげん": "語源",
    "カタ":   "カタカナ",
    "セット": "セットフレーズ",
    # leave "ゴロ" / "語源" / "カタカナ" / "セットフレーズ" / "イメージ" untouched
    # "例" and "熟語" are kept as-is for client confirmation.
}

def _cell_to_runs(cell_value):
    """Convert an openpyxl cell value (either str or CellRichText) into a
    list of {text, bold} runs. Plain strings ⇒ single non-bold run."""
    if cell_value is None:
        return []
    if isinstance(cell_value, CellRichText):
        runs = []
        for blk in cell_value:
            # blk is either a plain str (no formatting) or a TextBlock with .font.
            if hasattr(blk, "font") and blk.font is not None and blk.text:
                runs.append({"text": blk.text, "bold": bool(blk.font.b)})
            elif isinstance(blk, str) and blk:
                runs.append({"text": blk, "bold": False})
        return runs
    # Plain string
    s = str(cell_value)
    if not s:
        return []
    return [{"text": s, "bold": False}]


def _runs_text(runs):
    """Flatten runs back to plain text — used for warnings + indexing helpers."""
    return "".join(r["text"] for r in runs)


def _split_runs_by_separator(runs, sep_pattern):
    """Split a run list by a separator regex. Returns a list of sub-run lists.
    Used to split the combined "ゴロ／語源" format on '／'."""
    chunks = [[]]
    for r in runs:
        parts = re.split(sep_pattern, r["text"])
        for i, p in enumerate(parts):
            if i > 0:
                chunks.append([])
            if p:
                chunks[-1].append({"text": p, "bold": r["bold"]})
    # Drop empty chunks
    return [c for c in chunks if c and any(p["text"] for p in c)]


def parse_mnemonics(type_field: str, text_runs):
    """Return list of {type, runs:[{text,bold},...]}.

    type_field is plain string (e.g. 'ゴロ' or 'ゴロ・語源').
    text_runs is the rich-text run list from column G."""
    if not type_field and not text_runs:
        return []
    types = [t.strip() for t in re.split(r"[・･]", (type_field or "")) if t.strip()]
    types = [MNEMONIC_NORMALIZE.get(t, t) for t in types]
    # Split runs by '／' for multi-type entries.
    chunks = _split_runs_by_separator(text_runs, r"[／/]")
    # If no separator found, chunks == [runs].
    if not chunks:
        chunks = [text_runs] if text_runs else []
    out = []
    n = max(len(types), len(chunks))
    for i in range(n):
        runs_i = chunks[i] if i < len(chunks) else []
        type_i = types[i] if i < len(types) else (types[0] if types else "")
        # Strip leading/trailing whitespace on the runs as a whole.
        if runs_i:
            runs_i[0] = {"text": runs_i[0]["text"].lstrip(), "bold": runs_i[0]["bold"]}
            runs_i[-1] = {"text": runs_i[-1]["text"].rstrip(), "bold": runs_i[-1]["bold"]}
            runs_i = [r for r in runs_i if r["text"]]
        out.append({
            "type": type_i,
            "runs": runs_i,
        })
    return out

def block_of(id_int: int) -> int:
    if id_int <= 1100:
        return math.ceil(id_int / 48)
    return 23 + math.ceil((id_int - 1100) / 48)

def vol_of(id_int: int) -> int:
    return 1 if id_int <= 1100 else 2

def _parse_sheet(ws, source_label: str, warnings: list[str]) -> list[dict]:
    """Parse one sheet (9-col schema). Returns a list of record dicts.
    Skips header row and any blank rows. The schema is identical for vol.1
    and vol.2 since 2026-06-13 (落合 review, vol.2 例文 added).

    NOTE: reads with `values_only=False` so that the mnemonic-text column G
    can carry character-level bold formatting (CellRichText) — the spec
    requires the pronunciation-linked portion to be bold.

    Columns (vol.1 + vol.2 as of 2026-06-30 client update):
        A=B  B=見出し№  C=英単語  D=品詞  E=意味  F=覚え方  G=覚え方の具体的方法
        H=例文1  I=日本語訳1  J=例文2  K=日本語訳2"""
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if not row or row[1].value is None:
            continue
        try:
            wid = int(str(row[1].value).strip())
        except (TypeError, ValueError):
            continue
        def _s(v):
            return str(v).strip() if v is not None else ""
        word = _s(row[2].value)
        pos_raw = _s(row[3].value)
        meaning_raw = _s(row[4].value)
        mnemo_type_raw = _s(row[5].value)
        # Preserve rich-text bold for the mnemonic text column.
        mnemo_text_runs = _cell_to_runs(row[6].value)
        mnemo_text_raw = _runs_text(mnemo_text_runs)
        # Examples: column H/I = 例文1/日本語訳1, J/K = 例文2/日本語訳2 (added 2026-06-30).
        examples = []
        for en_col, ja_col in [(7, 8), (9, 10)]:
            en = _s(row[en_col].value) if len(row) > en_col else ""
            ja = _s(row[ja_col].value) if len(row) > ja_col else ""
            if en or ja:
                examples.append({"en": en, "ja": ja})

        pos_list, mode = parse_pos(pos_raw)
        meanings = split_meanings(meaning_raw, mode, pos_count=len(pos_list))
        mnemonics = parse_mnemonics(mnemo_type_raw, mnemo_text_runs)

        if not word:
            warnings.append(f"{source_label} row {row_idx} id={wid}: empty word")
        if not meaning_raw:
            warnings.append(f"{source_label} row {row_idx} id={wid}: empty meaning")
        if mode != "single" and len(meanings) < 2:
            warnings.append(
                f"{source_label} row {row_idx} id={wid}: "
                f"pos='{pos_raw}' suggests 2 meanings but parsed {len(meanings)}")

        records.append({
            "id": wid,
            "vol": vol_of(wid),
            "block": block_of(wid),
            "word": word,
            "pos_raw": pos_raw,
            "pos_list": pos_list,
            "meaning_mode": mode,
            "meanings": meanings,
            "mnemonics": mnemonics,
            # Back-compat single fields + new examples array.
            "example_en": examples[0]["en"] if examples else "",
            "example_ja": examples[0]["ja"] if examples else "",
            "examples": examples,
            "image_filename": f"{wid:04d}.webp",
            "pronunciation_hint": None,
        })
    return records


def _pick_sheet(wb, prefer_titles: list[str]):
    """Pick a worksheet, preferring known title prefixes, else the first
    sheet with > 1 row of content."""
    for t in prefer_titles:
        if t in wb.sheetnames:
            return wb[t]
    # Fallback: first non-empty sheet.
    for ws in wb.worksheets:
        if ws.max_row > 1:
            return ws
    return wb.worksheets[0]


def main():
    sources: list[tuple[str, Path]] = []
    vol1 = _pick_vol1_source()
    sources.append(("vol.1", vol1))
    vol2 = _pick_vol2_source()
    if vol2 is not None:
        sources.append(("vol.2", vol2))
    else:
        print("(no vol.2 master found — vol.1 only)")

    all_records: list[dict] = []
    warnings: list[str] = []
    used_files: list[str] = []
    for label, path in sources:
        print(f"{label}: {path.name}")
        used_files.append(path.name)
        # rich_text=True preserves character-level bold for mnemonic text.
        wb = openpyxl.load_workbook(path, data_only=True, rich_text=True)
        prefer = ["快単vol.1"] if label == "vol.1" else ["快単vol.2"]
        ws = _pick_sheet(wb, prefer)
        recs = _parse_sheet(ws, label, warnings)
        # Sanity: vol.1 records should be id 1..1100, vol.2 1101..2201.
        if label == "vol.1":
            recs = [r for r in recs if 1 <= r["id"] <= 1100]
        else:
            recs = [r for r in recs if 1101 <= r["id"] <= 2201]
        all_records += recs
        print(f"  parsed {len(recs)} records")

    # Sort by id and de-dup (last-write-wins; vol.2 file shouldn't overlap
    # vol.1 but we de-dup defensively).
    by_id: dict[int, dict] = {}
    for r in all_records:
        by_id[r["id"]] = r
    records = [by_id[i] for i in sorted(by_id)]

    # ── Merge pronunciation overrides ────────────────────────────────
    overrides_path = Path(__file__).resolve().parent / "pronunciation_overrides.json"
    pron_applied = 0
    if overrides_path.exists():
        with overrides_path.open(encoding="utf-8") as fp:
            doc = json.load(fp)
        for sid, entry in (doc.get("overrides") or {}).items():
            if not sid.isdigit():
                continue
            wid = int(sid)
            if not isinstance(entry, dict):
                continue
            hint = entry.get("hint")
            if hint and wid in by_id:
                by_id[wid]["pronunciation_hint"] = hint
                pron_applied += 1
        print(f"\npronunciation overrides applied: {pron_applied}")
    else:
        print("\n(no pronunciation_overrides.json found — skipped)")

    # Stats
    by_block: dict[int, int] = {}
    by_vol: dict[int, int] = {}
    by_mode = {"single": 0, "both_required": 0, "either_ok": 0}
    pos_seen: set[str] = set()
    mnemo_types: set[str] = set()
    for r in records:
        by_block[r["block"]] = by_block.get(r["block"], 0) + 1
        by_vol[r["vol"]] = by_vol.get(r["vol"], 0) + 1
        by_mode[r["meaning_mode"]] += 1
        for p in r["pos_list"]:
            pos_seen.add(p)
        for m in r["mnemonics"]:
            if m["type"]:
                mnemo_types.add(m["type"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as fp:
        json.dump({
            "schema_version": 2,           # bumped: now contains both vols
            "sources": used_files,
            "count": len(records),
            "stats": {
                "vol_counts": by_vol,
                "blocks_present": sorted(by_block.keys()),
                "block_word_counts": {str(k): by_block[k] for k in sorted(by_block)},
                "meaning_mode_counts": by_mode,
                "pos_seen": sorted(pos_seen),
                "mnemonic_types": sorted(mnemo_types),
                "warnings_count": len(warnings),
            },
            "words": records,
        }, fp, ensure_ascii=False, indent=2)

    print(f"\nwrote {OUT.relative_to(ROOT)}  records={len(records)}")
    print(f"vol counts: {by_vol}")
    print(f"blocks: {min(by_block)}..{max(by_block)}  total blocks={len(by_block)}")
    print(f"meaning modes: {by_mode}")
    print(f"POS seen: {sorted(pos_seen)}")
    print(f"mnemonic types: {sorted(mnemo_types)}")
    print(f"warnings: {len(warnings)} (first 8)")
    for w in warnings[:8]:
        print("  -", w)


if __name__ == "__main__":
    main()
