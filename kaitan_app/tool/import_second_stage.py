"""
Second Stage Excel -> JSON importer for 快単 (Kaitan).

Reads the client's two Second Stage master files:
  - Appli開発［foxgold共有］/アプリ基本データ Second Stage V1.xlsx  (vol.1, blocks 1-23)
  - Appli開発［foxgold共有］/アプリ基本データ Second Stage V2.xlsx  (vol.2, blocks 24-46)

Each row (columns A-J):
  A: B          (block number, filled only at first row of each block)
  B: 見出し№   (parent Phase 1 word id, str like "0097")
  C: 英単語     (parent word, informational)
  D: 品詞       (parent POS)
  E: 意味       (parent meaning)
  F: 問題       (relation code — free-form; NOT enum-restricted)
  G: 解答       (answer word/phrase/marker)
  H: 意味等     (answer meaning, if applicable)
  I: Gの音声    (TTS enable flag: True/False/blank)
  J: メモ       (implementation notes, mostly empty)

Design notes (v1.1 realization):
- Real-world data has hundreds of unique free-form relation strings beyond
  the base 15 codes (類/反/名/形/etc.). Do NOT enum them — keep the raw
  string in `relation`.
- Block/word_id are inherited from the last non-empty cell above (Excel
  merged-cell convention). A row with just a relation+answer belongs to
  the most recent word_id.
- TTS flag normalization: 'True' / 'TRUE' / ' FALSE ' → bool.

Output:
  kaitan_app/assets/content/second_stage.json
"""

from __future__ import annotations
import json
import sys
import re
from pathlib import Path
from collections import Counter

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT / "Appli開発［foxgold共有］"
OUT = Path(__file__).resolve().parents[1] / "assets" / "content" / "second_stage.json"
WORDS_JSON = Path(__file__).resolve().parents[1] / "assets" / "content" / "words.json"


def _bool(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().upper()
    return s == "TRUE"


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _pick(patterns: list[str]) -> Path | None:
    cands: list[Path] = []
    for p in patterns:
        cands += list(SRC_DIR.glob(p))
    if not cands:
        return None
    return max(cands, key=lambda p: p.stat().st_mtime)


def parse_ss_sheet(ws, source_label: str, warnings: list[str]) -> list[dict]:
    """Parse an SS sheet. Returns list of raw entry dicts (no id yet)."""
    entries: list[dict] = []
    current_block: int | None = None
    current_word_id: int | None = None
    current_word: str | None = None
    current_pos: str | None = None
    current_meaning: str | None = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        b_cell = row[0] if len(row) > 0 else None
        wid_cell = row[1] if len(row) > 1 else None
        w_cell = row[2] if len(row) > 2 else None
        pos_cell = row[3] if len(row) > 3 else None
        mean_cell = row[4] if len(row) > 4 else None
        rel_cell = row[5] if len(row) > 5 else None
        ans_cell = row[6] if len(row) > 6 else None
        ans_mean_cell = row[7] if len(row) > 7 else None
        tts_cell = row[8] if len(row) > 8 else None
        note_cell = row[9] if len(row) > 9 else None

        if b_cell is not None and str(b_cell).strip():
            try:
                current_block = int(str(b_cell).strip())
            except ValueError:
                pass

        if wid_cell is not None and str(wid_cell).strip():
            try:
                current_word_id = int(str(wid_cell).strip())
            except ValueError:
                pass
            current_word = _s(w_cell)
            current_pos = _s(pos_cell)
            current_meaning = _s(mean_cell)

        rel = _s(rel_cell)
        ans = _s(ans_cell)

        # Skip rows that have neither a relation nor an answer.
        if rel is None and ans is None:
            continue

        # Sanity gates
        if current_word_id is None:
            warnings.append(
                f"{source_label} row {row_idx}: entry without parent word_id (skipped): "
                f"rel={rel!r} ans={ans!r}"
            )
            continue
        if current_block is None:
            warnings.append(
                f"{source_label} row {row_idx} word_id={current_word_id}: entry without block"
            )

        entries.append({
            "block": current_block,
            "word_id": current_word_id,
            "word": current_word,
            "pos": current_pos,
            "meaning": current_meaning,
            "relation": rel or "",
            "answer": ans or "",
            "answer_meaning": _s(ans_mean_cell),
            "tts_enabled": _bool(tts_cell),
            "notes": _s(note_cell),
        })
    return entries


def base_category_of(relation: str) -> str | None:
    """Return the base category code (類/反/名/形/…) if the free-form relation
    starts with a known base code, else None. Mirrors the Dart-side
    SsRelationCategory.categoryOf() logic."""
    codes = ["類", "反", "前", "熟", "活", "品", "法", "複", "同音", "セ", "意",
             "名", "形", "副", "動"]
    trimmed = relation.strip()
    for code in codes:
        if trimmed == code:
            return code
        if trimmed.startswith(code + " ") or trimmed.startswith(code + "　"):
            return code
        if re.match(rf"^{code}\d+$", trimmed):
            return code
        if re.match(rf"^{code}[（(]", trimmed):
            return code
    return None


def main() -> None:
    v1 = _pick(["*Second Stage V1*.xlsx", "*Second stage V1*.xlsx"])
    v2 = _pick(["*Second Stage V2*.xlsx", "*Second stage V2*.xlsx"])
    if v1 is None or v2 is None:
        print(f"Missing SS source file(s). v1={v1} v2={v2}", file=sys.stderr)
        sys.exit(1)

    words_doc = json.loads(WORDS_JSON.read_text(encoding="utf-8"))
    valid_word_ids = {w["id"] for w in words_doc["words"]}

    warnings: list[str] = []
    used_files: list[str] = []
    all_entries: list[dict] = []

    for label, path, prefer_sheet in [
        ("V1", v1, "快単vol.1"),
        ("V2", v2, "1101-1214"),
    ]:
        print(f"{label}: {path.name}")
        used_files.append(path.name)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[prefer_sheet] if prefer_sheet in wb.sheetnames else wb.active
        rows = parse_ss_sheet(ws, label, warnings)
        print(f"  parsed {len(rows)} entries from sheet '{ws.title}'")
        all_entries.extend(rows)

    # Assign a stable 1-based id after full ordering. Order by (word_id, then
    # position within file) so that entries for the same headword stay
    # contiguous — matters for the UI which shows them together.
    all_entries.sort(key=lambda e: (e["word_id"] or 0,))

    # Cross-check: every word_id must exist in Phase 1 words.json.
    orphans: list[int] = []
    for e in all_entries:
        if e["word_id"] not in valid_word_ids:
            orphans.append(e["word_id"])
    orphans = sorted(set(orphans))
    if orphans:
        warnings.append(f"orphan word_ids (not in words.json): {orphans[:20]}"
                        + (f" (total {len(orphans)})" if len(orphans) > 20 else ""))

    entries_out: list[dict] = []
    for idx, e in enumerate(all_entries, start=1):
        entries_out.append({
            "id": idx,
            "word_id": e["word_id"],
            "block": e["block"],
            "relation": e["relation"],
            "base_category": base_category_of(e["relation"]),
            "answer": e["answer"],
            "answer_meaning": e["answer_meaning"],
            "tts_enabled": e["tts_enabled"],
            "notes": e["notes"],
        })

    # Stats
    by_block = Counter(e["block"] for e in entries_out if e["block"])
    by_cat = Counter(e["base_category"] or "(free-form)" for e in entries_out)
    with_tts = sum(1 for e in entries_out if e["tts_enabled"])
    with_meaning = sum(1 for e in entries_out if e["answer_meaning"])
    unique_words = len({e["word_id"] for e in entries_out})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "schema_version": 1,
        "sources": used_files,
        "count": len(entries_out),
        "stats": {
            "unique_word_ids": unique_words,
            "by_block": {str(k): v for k, v in sorted(by_block.items())},
            "by_base_category": dict(by_cat.most_common()),
            "with_tts_enabled": with_tts,
            "with_answer_meaning": with_meaning,
            "warnings_count": len(warnings),
        },
        "entries": entries_out,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print()
    print(f"wrote {OUT.relative_to(ROOT)}  entries={len(entries_out)}")
    print(f"unique word_ids: {unique_words} (of {len(valid_word_ids)} Phase 1 words)")
    print(f"blocks: {min(by_block)}..{max(by_block)} (n={len(by_block)})")
    print(f"tts_enabled: {with_tts}, with_meaning: {with_meaning}")
    print(f"top base categories: {list(by_cat.most_common(8))}")
    print(f"warnings: {len(warnings)}")
    for w in warnings[:8]:
        print(f"  - {w}")


if __name__ == "__main__":
    main()
